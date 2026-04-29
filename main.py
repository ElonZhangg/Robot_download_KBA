# -*- coding: utf-8 -*-
"""
主程序 - 自动化登录、搜索、下载PDF
"""
import sys
import time
import tkinter as tk
from tkinter import simpledialog, messagebox

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchWindowException
import openpyxl

# 导入自定义模块
import config
from locator_helper import LocatorHelper
from download_monitor import DownloadMonitor

def setup_driver():
    """配置Chrome浏览器，设置下载目录"""
    chrome_options = webdriver.ChromeOptions()
    # 设置下载路径
    prefs = {
        "download.default_directory": config.DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    chrome_options.add_experimental_option("prefs", prefs)
    # 禁用自动化提示（可选）
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    if config.CHROME_DRIVER_PATH:
        driver = webdriver.Chrome(executable_path=config.CHROME_DRIVER_PATH, options=chrome_options)
    else:
        driver = webdriver.Chrome(options=chrome_options)
    driver.maximize_window()
    return driver

def wait_for_user_input(prompt, timeout=None):
    """弹窗等待用户输入（用于用户名和密码的手动输入）"""
    root = tk.Tk()
    root.withdraw()
    user_input = simpledialog.askstring("输入", prompt, parent=root)
    root.destroy()
    return user_input

def login(driver, locator):
    """登录流程：等待管理员输入用户名和密码，然后点击登录按钮"""
    print("正在打开登录页面...")
    driver.get(config.TARGET_URL)
    time.sleep(config.WAIT_AFTER_OPEN)
    
    # 1. 定位用户名输入框（支持手动回退）
    username_input = locator.find_element("username", config.LOGIN_USERNAME_SELECTOR, timeout=10)
    # 等待管理员输入用户名（弹出对话框，用户手动在浏览器中输入）
    # 为了符合“等待管理员输入”，程序弹窗提示，用户确认后继续
    msg = "请在浏览器中手动输入用户名（elon.zhang@emerson.com），然后点击确定"
    messagebox.showinfo("等待用户名", msg)
    # 等待7秒（用户输入时间）
    time.sleep(config.WAIT_AFTER_USERNAME)
    
    # 2. 密码输入框
    password_input = locator.find_element("password", config.LOGIN_PASSWORD_SELECTOR, timeout=10)
    msg = "请在浏览器中手动输入密码（12345678），然后点击确定"
    messagebox.showinfo("等待密码", msg)
    time.sleep(config.WAIT_AFTER_PASSWORD)
    
    # 3. 点击登录按钮
    login_btn = locator.find_element("login_button", config.LOGIN_BUTTON_SELECTOR, timeout=10)
    login_btn.click()
    print("登录按钮已点击，等待页面跳转...")
    time.sleep(3)  # 等待登录完成

def get_article_ids_from_excel(excel_path):
    """从Excel文件的E列读取Article ID（从第2行开始，直到空值）"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    article_ids = []
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=5).value  # E列
        if cell_value is None or str(cell_value).strip() == "":
            break
        article_ids.append(str(cell_value).strip())
    wb.close()
    print(f"共读取到 {len(article_ids)} 个Article ID")
    return article_ids

def search_and_download(driver, locator, article_id, download_monitor):
    """
    对单个Article ID执行搜索、打开详情、下载PDF
    返回是否成功
    """
    try:
        # ---------- 步骤7：输入Article ID并搜索 ----------
        search_input = locator.find_element("search_input", config.SEARCH_INPUT_SELECTOR, timeout=10)
        search_input.clear()
        search_input.send_keys(article_id)
        search_input.send_keys("\n")  # 回车搜索
        print(f"已搜索 Article ID: {article_id}")
        
        # ---------- 步骤8：等待10秒 ----------
        time.sleep(config.WAIT_AFTER_SEARCH)
        
        # ---------- 步骤9：双击第一条搜索记录 ----------
        first_result = locator.find_element("first_result", config.FIRST_RESULT_SELECTOR, timeout=10)
        # 获取当前窗口句柄
        main_window = driver.current_window_handle
        # 双击
        ActionChains(driver).double_click(first_result).perform()
        print("已双击第一条记录")
        
        # ---------- 步骤10：等待7秒，切换到新打开的标签页 ----------
        time.sleep(config.WAIT_AFTER_OPEN_DETAIL)
        # 获取所有窗口句柄，切换到最新打开的（通常是最后一个）
        all_handles = driver.window_handles
        new_window = [h for h in all_handles if h != main_window][0]
        driver.switch_to.window(new_window)
        print("已切换到详情页标签页")
        
        # ---------- 步骤11：双击右上角download图标 ----------
        download_icon = locator.find_element("download_icon", config.DOWNLOAD_ICON_SELECTOR, timeout=10)
        ActionChains(driver).double_click(download_icon).perform()
        print("已双击下载图标")
        
        # ---------- 步骤12：等待下载完成 ----------
        time.sleep(config.WAIT_AFTER_DOWNLOAD)
        downloaded_file = download_monitor.wait_for_download(timeout=config.DOWNLOAD_TIMEOUT)
        print(f"PDF下载完成：{downloaded_file}")
        
        # ---------- 步骤14：关闭详情页标签页 ----------
        driver.close()
        driver.switch_to.window(main_window)
        print("已关闭详情页，返回搜索结果页")
        return True
        
    except Exception as e:
        print(f"处理 Article ID {article_id} 时出错：{e}")
        # 尝试恢复到主窗口（如果可能）
        try:
            if len(driver.window_handles) > 1:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
        except:
            pass
        return False

def main():
    print("=== 自动化程序启动 ===")
    # 初始化下载监控
    download_monitor = DownloadMonitor(config.DOWNLOAD_DIR)
    # 清空下载目录（可选，避免残留文件干扰）
    # download_monitor.clear_old_files()
    
    # 启动浏览器
    driver = setup_driver()
    locator = LocatorHelper(driver)
    
    try:
        # 登录
        login(driver, locator)
        
        # 获取Article ID列表
        article_ids = get_article_ids_from_excel(config.EXCEL_PATH)
        if not article_ids:
            print("Excel中未找到任何Article ID，请检查E列内容")
            return
        
        # 循环处理每个ID
        for idx, aid in enumerate(article_ids, start=1):
            print(f"\n--- 处理第 {idx}/{len(article_ids)} 个：{aid} ---")
            success = search_and_download(driver, locator, aid, download_monitor)
            if not success:
                print(f"Article ID {aid} 处理失败，是否继续？")
                # 可选：弹窗询问是否继续
                root = tk.Tk()
                root.withdraw()
                retry = messagebox.askyesno("错误", f"处理 {aid} 失败，是否继续下一个？")
                root.destroy()
                if not retry:
                    break
            # 每处理完一个，等待一下
            time.sleep(2)
        
        print("\n所有任务执行完毕！")
        messagebox.showinfo("完成", "批量下载已完成！")
        
    except Exception as e:
        print(f"程序发生严重错误：{e}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("错误", f"程序异常：{e}")
    finally:
        # 保持浏览器打开30秒后关闭（方便查看结果）
        print("程序将在30秒后关闭浏览器...")
        time.sleep(30)
        driver.quit()

if __name__ == "__main__":
    main()